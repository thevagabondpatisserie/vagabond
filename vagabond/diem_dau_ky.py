"""Nap diem dau ky tu tep Excel, va chay lai diem khi doi ty le tich.

Hai viec rieng biet, de chung mot tep vi cung dung vao so `Vagabond So Diem`
va cung phai tuan mot luat: KHONG SUA BUT CU (QT-20). Muon doi mot con so
da vao so thi ghi mot but nguoc, khong tay dong cu.

Ca hai ham deu CHAY THU truoc. Goi khong tham so la chi doc va bao cao,
khong ghi mot dong nao. Phai truyen chay_that=1 moi ghi. Lam vay vi day la
hai thao tac dong vao du lieu QUA KHU cua hang tram khach: xem bao cao truoc
roi moi quyet dinh thi con duong lui.
"""

import io
from contextlib import ExitStack

import frappe
from frappe.utils import cint, flt, now_datetime
from frappe.utils.synchronization import LockTimeoutError, filelock

from vagabond.lib import sdt_so

# Ten khoa dung chung cho MOI viec ghi hang loat vao so diem.
#
# Vi sao phai co - bat duoc bang mau chinh minh, 16/08/2026
# --------------------------------------------------------
# Lan chay that dau tien cua nap_dau_ky() bi goi HAI LAN gan nhu cung luc
# (yeu cau dau treo qua lau nen ben goi thu lai). Hai luot cung doc bang
# "khach da co but dau ky" khi CHUA ben nao commit, nen ca hai deu thay
# bang rong va deu ghi. Ket qua: 16.083 khach moi nguoi hai but, quy diem
# thanh 943 trieu thay vi 471 trieu.
#
# Kiem "da co chua" TRONG ma khong bao gio chan duoc chuyen nay, phai chan
# tu ngoai bang khoa. Day dung la loi ma ban_hang._khoa_dong_bo da gap va
# da giai ngay 07/08/2026 voi hai hoa don cho mot don Pancake; le ra phai
# be nep do sang day tu dau.
#
# Khoa TEP chu khong phai khoa bo nho dem: tien trinh chet thi he dieu
# hanh tu tha khoa, con khoa bo nho dem chet giua chung se de lai chia
# khoa mo coi chan sach moi duong ghi so.
KHOA_SO_DIEM = "vgb_ghi_so_diem"


def _khoa(cho=5):
	"""Chi cho MOT viec ghi hang loat vao so diem tai mot thoi diem."""
	pila = ExitStack()
	try:
		pila.enter_context(filelock(KHOA_SO_DIEM, timeout=cho))
	except LockTimeoutError:
		pila.close()
		frappe.throw(
			"Máy đang chạy dở một lượt ghi sổ điểm. Chờ cho xong rồi hãy bấm lại, "
			"đừng bấm thêm lần nữa - bấm hai lần là điểm bị cộng đôi."
		)
	return pila

SO_DIEM = "Vagabond So Diem"
LOAI_DAU_KY = "So du dau ky"
LOAI_DOI_TY_LE = "Dieu chinh ty le tich diem"
LOAI_TICH = "Tich tu hoa don"

# Ten cot co the gap trong tep. Doc theo TEN chu khong theo vi tri: tep do
# nguoi lam tay, hom nay cot G hom sau cot H, khoa cung vi tri la sai am tham.
COT_DIEM = ("tich diem", "diem", "diem tich luy", "so diem", "diem hien co")
COT_MA = ("ma khach", "ma khach hang", "customer", "ma kh", "id")
COT_SDT = ("dien thoai", "so dien thoai", "sdt", "mobile", "phone")
COT_TEN = ("ten khach", "ten khach hang", "ho ten", "customer name", "ten")


def _khong_dau(s):
	"""Bo dau tieng Viet va ha chu thuong, de so ten cot cho de."""
	import unicodedata

	x = unicodedata.normalize("NFD", str(s or ""))
	x = "".join(ch for ch in x if unicodedata.category(ch) != "Mn")
	return x.replace("đ", "d").replace("Đ", "d").strip().lower()


def _tim_cot(dau_bang, ten_co_the):
	"""Chi so cot dau tien khop mot trong cac ten. -1 la khong thay."""
	sach = [_khong_dau(x) for x in dau_bang]
	for muon in ten_co_the:
		for i, co in enumerate(sach):
			if co == muon:
				return i
	# Chua khop tuyet doi thi cho khop chua dung, nhung phai bao ra.
	for muon in ten_co_the:
		for i, co in enumerate(sach):
			if co and muon in co:
				return i
	return -1


def _doc_bang(noi_dung, ten_tep=""):
	"""Tra (dau_bang, cac_dong). Nhan ca .xlsx lan .csv.

	Phai nhan ca hai vi anh Viet gui khi thi tep Excel khi thi tep CSV xuat
	tu cung mot man hinh, va hai ban do KHONG giong nhau - xem tai lieu
	ngay 16/08/2026.
	"""
	if str(ten_tep or "").lower().endswith(".csv") or noi_dung[:1] in (b"\xef", b"S", b"s"):
		try:
			return _doc_csv(noi_dung)
		except Exception:
			pass
	return _doc_excel(noi_dung)


def _doc_csv(noi_dung):
	import csv

	# utf-8-sig: tep xuat tu Excel gan nhu luon co BOM o dau, khong bo thi
	# ten cot dau tien mang ba ky tu rac va khong khop voi bat ky ten nao.
	chu = noi_dung.decode("utf-8-sig", errors="replace")
	doc = list(csv.reader(io.StringIO(chu)))
	if not doc:
		frappe.throw("Tệp CSV không có dòng nào. Kiểm tra lại tệp giúp em.")
	for i, r in enumerate(doc[:15]):
		if _tim_cot(r, COT_DIEM) >= 0:
			return r, doc[i + 1 :]
	return doc[0], doc[1:]


def _doc_excel(noi_dung):
	"""Tra (dau_bang, cac_dong). Chi doc trang dau tien."""
	try:
		from openpyxl import load_workbook
	except ImportError:
		frappe.throw("Máy chủ chưa có thư viện đọc Excel. Báo em để cài openpyxl.")

	wb = load_workbook(io.BytesIO(noi_dung), read_only=True, data_only=True)
	ws = wb[wb.sheetnames[0]]
	dong = []
	for r in ws.iter_rows(values_only=True):
		dong.append(list(r))
	wb.close()
	if not dong:
		frappe.throw("Tệp Excel không có dòng nào. Kiểm tra lại tệp giúp em.")

	# Dau bang co the khong nam o dong 1: nhieu tep co dong tieu de, dong
	# trong, roi moi toi ten cot. Tim dong dau tien co tu "diem" trong do.
	for i, r in enumerate(dong[:15]):
		if _tim_cot(r, COT_DIEM) >= 0:
			return r, dong[i + 1 :]
	return dong[0], dong[1:]


def _so_nguyen(v):
	"""Doc mot o thanh so nguyen. Doc khong ra thi tra None chu khong tra 0.

	Tra 0 thi mot o hong se thanh "khach nay 0 diem" va khong ai biet; tra
	None thi no vao muc bo qua trong bao cao va nguoi doc nhin thay.
	"""
	if v is None:
		return None
	if isinstance(v, (int, float)):
		return int(round(v))
	x = str(v).strip().replace(".", "").replace(",", "").replace(" ", "")
	if not x or not x.lstrip("-").isdigit():
		return None
	return int(x)


def _ban_do_sdt():
	"""So dien thoai -> ma khach, dung cho tep khong co cot ma khach."""
	ra = {}
	rows = frappe.db.sql(
		"""
		select dl.link_name ma, c.mobile_no, c.phone
		from `tabContact` c
		join `tabDynamic Link` dl on dl.parent = c.name and dl.parenttype = 'Contact'
		where dl.link_doctype = 'Customer'
		""",
		as_dict=True,
	)
	for r in rows:
		for s in (r.get("mobile_no"), r.get("phone")):
			k = sdt_so(s)
			# So trung o hai khach thi BO CA HAI: gan diem cho nham nguoi
			# con te hon la khong gan.
			if not k:
				continue
			if k in ra and ra[k] != r["ma"]:
				ra[k] = None
			elif k not in ra:
				ra[k] = r["ma"]
	return {k: v for k, v in ra.items() if v}


@frappe.whitelist()
def nap_dau_ky(file_url=None, chay_that=0, ngay=None):
	"""Doc cot Tich diem tu tep Excel, ghi but "So du dau ky" cho tung khach.

	file_url: duong dan tep da tai len he (dang /private/files/... hoac
	/files/...). Lay bang cach keo tep vao mot ban ghi bat ky roi copy
	duong dan, hoac tai len o Trinh quan ly tep.

	Chay khong co chay_that thi CHI DOC VA BAO CAO.
	"""
	_chi_quan_ly()
	if not (file_url or "").strip():
		frappe.throw(
			"Chưa có tệp. Tải \"Khách hàng tổng hợp.xlsx\" lên Trình quản lý tệp "
			"rồi truyền đường dẫn tệp vào giúp em."
		)
	noi_dung = _doc_tep(file_url)
	dau, dong = _doc_bang(noi_dung, file_url)

	i_diem = _tim_cot(dau, COT_DIEM)
	i_ma = _tim_cot(dau, COT_MA)
	i_sdt = _tim_cot(dau, COT_SDT)
	i_ten = _tim_cot(dau, COT_TEN)
	if i_diem < 0:
		frappe.throw(
			"Không thấy cột điểm trong tệp. Các cột đọc được: %s. Đổi tên cột "
			"thành \"Tích điểm\" rồi tải lại giúp em." % ", ".join(str(x) for x in dau if x)
		)
	if i_ma < 0 and i_sdt < 0:
		frappe.throw(
			"Tệp phải có cột mã khách hoặc cột số điện thoại thì em mới biết "
			"cộng điểm cho ai. Các cột đọc được: %s." % ", ".join(str(x) for x in dau if x)
		)

	theo_sdt = _ban_do_sdt() if i_sdt >= 0 else {}
	# Khach da co but dau ky roi thi BO QUA. Chay ham nay lan hai khong
	# duoc cong doi diem cua ai.
	da_co = set(
		frappe.get_all(SO_DIEM, filters={"loai": LOAI_DAU_KY}, pluck="khach", limit_page_length=0)
	)

	ra = {
		"chay_that": cint(chay_that),
		"tong_dong": len(dong),
		"se_ghi": [],
		"bo_qua_da_co": 0,
		"bo_qua_khong_diem": 0,
		"bo_qua_khong_thay_khach": [],
		"bo_qua_o_hong": [],
		"tong_diem": 0,
	}

	for n, r in enumerate(dong, start=2):
		if not r or all(x is None or str(x).strip() == "" for x in r):
			continue
		diem = _so_nguyen(r[i_diem] if i_diem < len(r) else None)
		if diem is None:
			o = r[i_diem] if i_diem < len(r) else ""
			if str(o or "").strip():
				ra["bo_qua_o_hong"].append({"dong": n, "gia_tri": str(o)[:40]})
			continue
		if diem <= 0:
			ra["bo_qua_khong_diem"] += 1
			continue

		ma = ""
		if i_ma >= 0 and i_ma < len(r):
			ma = str(r[i_ma] or "").strip()
		if ma and not frappe.db.exists("Customer", ma):
			ma = ""
		if not ma and i_sdt >= 0 and i_sdt < len(r):
			ma = theo_sdt.get(sdt_so(r[i_sdt]), "") or ""
		if not ma:
			nhan = ""
			if i_ten >= 0 and i_ten < len(r):
				nhan = str(r[i_ten] or "")[:40]
			ra["bo_qua_khong_thay_khach"].append({"dong": n, "ten": nhan, "diem": diem})
			continue
		if ma in da_co:
			ra["bo_qua_da_co"] += 1
			continue

		da_co.add(ma)
		ra["se_ghi"].append({"khach": ma, "diem": diem})
		ra["tong_diem"] += diem

	ra["so_khach"] = len(ra["se_ghi"])
	# Canh bao khi CA COT bang 0.
	#
	# Bat duoc 16/08/2026: ban CSV xuat cung ngay co cot Tich diem bang 0 o
	# ca 56.960 dong, trong khi ban .xlsx xuat truoc do co 16.229 khach mang
	# tong 472 trieu diem. Chay im lang tren ban CSV thi ghi duoc 0 but va
	# bao "xong", nguoi doc tuong da nap du.
	if not ra["se_ghi"] and ra["bo_qua_khong_diem"] > 100:
		ra["canh_bao"] = (
			"Cột điểm của tệp này bằng 0 ở toàn bộ %d dòng đọc được, nên không có "
			"gì để nạp. Nhiều khả năng đây là bản xuất thiếu cột điểm. Kiểm tra "
			"lại tệp nguồn trước khi chạy thật." % ra["bo_qua_khong_diem"]
		)
	# Bao cao dai thi cat bot phan liet ke, nhung GIU nguyen con so dem:
	# cat im lang la nguoi doc tuong da phu het.
	ra["bo_qua_khong_thay_khach_tong"] = len(ra["bo_qua_khong_thay_khach"])
	ra["bo_qua_khong_thay_khach"] = ra["bo_qua_khong_thay_khach"][:50]
	ra["bo_qua_o_hong_tong"] = len(ra["bo_qua_o_hong"])
	ra["bo_qua_o_hong"] = ra["bo_qua_o_hong"][:50]

	if not cint(chay_that):
		ra["se_ghi"] = ra["se_ghi"][:50]
		ra["ghi_chu"] = "Đây là bản chạy thử, chưa ghi gì vào sổ. Xem kỹ rồi gọi lại với chay_that=1."
		return ra

	from vagabond.khach_hang import _ghi_so_diem

	luc = ngay or now_datetime()
	da_ghi = 0
	pila = _khoa()
	# Doc LAI danh sach da co NGAY SAU khi cam khoa. Luot thu hai vao toi
	# day se thay luot thu nhat da ghi xong va bo qua sach, thay vi ghi de
	# len mot lan nua.
	da_co_2 = set(
		frappe.get_all(SO_DIEM, filters={"loai": LOAI_DAU_KY}, pluck="khach", limit_page_length=0)
	)
	ra["se_ghi"] = [x for x in ra["se_ghi"] if x["khach"] not in da_co_2]
	ra["bo_qua_da_co"] += len(da_co_2)
	for x in ra["se_ghi"]:
		try:
			_ghi_so_diem(
				x["khach"],
				x["diem"],
				LOAI_DAU_KY,
				None,
				"Số dư đầu kỳ nạp từ tệp Khách hàng tổng hợp, ngày %s." % luc,
			)
			da_ghi += 1
		except Exception:
			frappe.log_error(frappe.get_traceback(), "diem_dau_ky: nap %s" % x["khach"])
	frappe.db.commit()
	pila.close()
	ra["da_ghi"] = da_ghi
	ra["se_ghi"] = ra["se_ghi"][:50]
	return ra


def _doc_tep(file_url):
	ten = frappe.db.get_value("File", {"file_url": file_url}, "name")
	if not ten:
		frappe.throw("Không tìm thấy tệp %s trên hệ. Tải lại tệp rồi thử lại giúp em." % file_url)
	return frappe.get_doc("File", ten).get_content()


# --------------------------------------------------- chay lai theo ty le moi


@frappe.whitelist()
def tinh_lai_theo_ty_le(chay_that=0, tu_ngay=None, den_ngay=None):
	"""Quet cac but "Tich tu hoa don" va bu chenh lech khi ty le hang doi.

	KHONG sua but cu (QT-20). Voi moi hoa don lech, ghi MOT but moi loai
	"Dieu chinh ty le tich diem" mang dung phan chenh. Doc so tu tren xuong
	van thay day du: da tich bao nhieu theo muc cu, roi dieu chinh bao nhieu,
	vi sao.

	Ty le lay tu bang hang HIEN TAI, va lay theo hang khach DANG deo. Neu
	khach len hang giua chung thi day la mot han che co that: hoa don cu
	se duoc tinh theo hang moi. Ham bao ra so hoa don roi vao truong hop do
	de nguoi doc biet, chu khong lang le lam.
	"""
	_chi_quan_ly()
	from vagabond.khach_hang import _ghi_so_diem, diem_cho_don

	loc = {"loai": LOAI_TICH}
	if tu_ngay:
		loc["ngay"] = [">=", tu_ngay]
	but = frappe.get_all(
		SO_DIEM,
		filters=loc,
		fields=["name", "khach", "hoa_don", "diem", "ngay"],
		limit_page_length=0,
		order_by="ngay asc",
	)
	if den_ngay:
		but = [b for b in but if str(b.get("ngay") or "") <= str(den_ngay) + " 23:59:59"]

	# Hoa don da co but dieu chinh roi thi bo qua: chay ham hai lan khong
	# duoc dieu chinh chong len nhau.
	da_chinh = set(
		frappe.get_all(
			SO_DIEM, filters={"loai": LOAI_DOI_TY_LE}, pluck="hoa_don", limit_page_length=0
		)
	)

	ty_le_hang = {
		d["name"]: flt(d["tich_diem"])
		for d in frappe.get_all("Vagabond Hang Khach", fields=["name", "tich_diem"], limit_page_length=0)
	}
	hang_khach = {}
	ra = {
		"chay_that": cint(chay_that),
		"tong_but": len(but),
		"khop": 0,
		"bo_qua_da_chinh": 0,
		"khong_doc_duoc_don": 0,
		"lech": [],
		"tong_chenh": 0,
	}

	for b in but:
		hd = b.get("hoa_don")
		if not hd:
			continue
		if hd in da_chinh:
			ra["bo_qua_da_chinh"] += 1
			continue
		tong = frappe.db.get_value("Sales Invoice", hd, "grand_total")
		if tong is None:
			ra["khong_doc_duoc_don"] += 1
			continue
		kh = b["khach"]
		if kh not in hang_khach:
			hang_khach[kh] = (frappe.db.get_value("Customer", kh, "vgb_hang") or "").strip()
		ty_le = ty_le_hang.get(hang_khach[kh], 0)
		dung = diem_cho_don(tong, ty_le)
		chenh = int(dung) - int(round(flt(b["diem"])))
		if chenh == 0:
			ra["khop"] += 1
			continue
		ra["lech"].append(
			{
				"hoa_don": hd,
				"khach": kh,
				"hang": hang_khach[kh],
				"da_tich": int(round(flt(b["diem"]))),
				"dung_ra": int(dung),
				"chenh": chenh,
			}
		)
		ra["tong_chenh"] += chenh

	ra["so_don_lech"] = len(ra["lech"])
	ra["so_khach_lech"] = len({x["khach"] for x in ra["lech"]})

	if not cint(chay_that):
		ra["lech"] = ra["lech"][:100]
		ra["ghi_chu"] = (
			"Đây là bản chạy thử, chưa ghi gì vào sổ. Tổng chênh âm nghĩa là khách "
			"BỊ TRỪ BỚT điểm đã có. Xem kỹ rồi gọi lại với chay_that=1."
		)
		return ra

	da_ghi = 0
	pila = _khoa()
	da_chinh_2 = set(
		frappe.get_all(SO_DIEM, filters={"loai": LOAI_DOI_TY_LE}, pluck="hoa_don", limit_page_length=0)
	)
	ra["lech"] = [x for x in ra["lech"] if x["hoa_don"] not in da_chinh_2]
	for x in ra["lech"]:
		try:
			_ghi_so_diem(
				x["khach"],
				x["chenh"],
				LOAI_DOI_TY_LE,
				x["hoa_don"],
				"Tính lại theo tỷ lệ hạng %s hiện hành: đã tích %d, đúng ra %d."
				% (x["hang"] or "(chưa xếp)", x["da_tich"], x["dung_ra"]),
			)
			da_ghi += 1
		except Exception:
			frappe.log_error(frappe.get_traceback(), "diem_dau_ky: tinh lai %s" % x["hoa_don"])
	frappe.db.commit()
	pila.close()
	ra["da_ghi"] = da_ghi
	ra["lech"] = ra["lech"][:100]
	return ra


@frappe.whitelist()
def go_but_dau_ky_trung(chay_that=0):
	"""Go cac but "So du dau ky" bi ghi TRUNG do luot chay doi 16/08/2026.

	Chi dong toi khach co TU HAI but dau ky tro len, va chi bo phan thua,
	giu lai but SOM NHAT. Khach chi co mot but thi ham nay khong cham vao.

	Vi sao xoa chu khong ghi but dao (QT-20)
	----------------------------------------
	QT-20 la de bao ve CHUNG TU: mot but do nghiep vu that sinh ra thi
	khong duoc xoa dau vet. But thua o day khong phai nghiep vu - no la
	ban sao do mot yeu cau bi goi hai lan trong cung mot phut, chua ai
	nhin thay, chua khach nao tieu vao. Ghi but dao thi so co ba dong cho
	mot su kien khong he xay ra, va nguoi doc so ba thang sau se khong
	hieu chuyen gi.

	Nhung day van la QUYET DINH CUA ANH VIET chu khong phai cua may: ham
	MAC DINH CHAY THU, va bao cao du de doi chieu truoc khi bam that.
	"""
	_chi_quan_ly()
	rows = frappe.db.sql(
		"""
		select name, khach, diem, creation
		from `tab%s` where loai = %%s order by khach, creation
		"""
		% SO_DIEM,
		(LOAI_DAU_KY,),
		as_dict=True,
	)
	theo_khach = {}
	for r in rows:
		theo_khach.setdefault(r["khach"], []).append(r)

	xoa, giu_diem, xoa_diem, lech = [], 0.0, 0.0, []
	for khach, ds in theo_khach.items():
		if len(ds) < 2:
			giu_diem += flt(ds[0]["diem"])
			continue
		# So tien cua cac ban sao phai BANG NHAU. Khong bang thi day khong
		# phai ban sao ma la hai su kien khac nhau - bo qua, bao ra de
		# nguoi xem, tuyet doi khong doan.
		if len({round(flt(x["diem"]), 4) for x in ds}) != 1:
			lech.append({"khach": khach, "diem": [flt(x["diem"]) for x in ds]})
			giu_diem += sum(flt(x["diem"]) for x in ds)
			continue
		giu_diem += flt(ds[0]["diem"])
		for x in ds[1:]:
			xoa.append(x["name"])
			xoa_diem += flt(x["diem"])

	ra = {
		"chay_that": cint(chay_that),
		"tong_but": len(rows),
		"so_khach": len(theo_khach),
		"se_xoa": len(xoa),
		"diem_se_bot": xoa_diem,
		"diem_con_lai": giu_diem,
		"khach_lech_khong_dong": lech[:50],
		"so_khach_lech": len(lech),
	}
	if not cint(chay_that):
		ra["ghi_chu"] = "Bản chạy thử, chưa xoá gì. Đối chiếu số điểm còn lại rồi mới chạy thật."
		return ra

	pila = _khoa()
	from vagabond.khach_hang import _tinh_lai_so_du

	da = 0
	for ten in xoa:
		try:
			frappe.delete_doc(SO_DIEM, ten, ignore_permissions=True, force=True, delete_permanently=True)
			da += 1
		except Exception:
			frappe.log_error(frappe.get_traceback(), "diem_dau_ky: go but trung %s" % ten)
	frappe.db.commit()
	# Tinh lai so du tren Customer cho tung khach vua dong toi: o vgb_diem
	# la ban tong hop, khong tinh lai thi man hinh con hien so cu.
	for khach in theo_khach:
		try:
			_tinh_lai_so_du(khach)
		except Exception:
			pass
	frappe.db.commit()
	pila.close()
	ra["da_xoa"] = da
	return ra


def _chi_quan_ly():
	if not ({"System Manager", "Accounts Manager"} & set(frappe.get_roles())):
		frappe.throw("Chỉ quản trị hệ thống hoặc kế toán trưởng chạy được việc này.")
