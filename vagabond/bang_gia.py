# -*- coding: utf-8 -*-
"""Bang gia mua nguyen vat lieu, lam tren app cho Uyen.

Anh Viet 14/08/2026 hoi em nen dat cho khai gia o dau, va em tra loi la
KHONG nen de trong Danh muc san pham. Ly do nam o su co lap xuong:

  Mon NVLT00354 "Lap xuong, Tui 400gr" co don vi kho la GRAM. Nguoi nhap
  nghi theo TUI, go 135.185 vao o don gia moi gram. Ket qua 2.700 g o kho
  Baker thanh 365 trieu dong thay vi 912 nghin.

Sai o day khong phai sai gia, ma la SAI DON VI. Neu chi lam mot o cho Uyen
go gia thi lan sau van go nham y het. Nen man nay bat khai ba thu:

  don vi mua       - Tui, Thung, Bao... thu Uyen doc tren hoa don NCC
  quy doi          - 1 tui bang bao nhieu don vi kho (1 tui = 400 gram)
  gia moi don vi mua - 135.185 d mot tui

May tu chia ra 338 d moi gram. Uyen khong bao gio phai go so per gram nua.

Ghi vao dau: mot ban ghi "Item Price" cua ERPNext tren bang gia mua, cong
voi mot dong "UOM Conversion Detail" tren chinh mat hang do. Hai thu nay
la duong ma don dat hang va gia von cua ERPNext van doc san - khong dung
bang rieng de roi phai tu noi lai.
"""

import base64
import io
import json

import frappe
from frappe.utils import cint, flt, nowdate

QUYEN_XEM = {
	"System Manager", "Purchase User", "Purchase Manager",
	"Accounts Manager", "Accounts User", "Bộ phận đặt hàng", "Stock Manager",
}
QUYEN_SUA = {
	"System Manager", "Purchase Manager", "Purchase User",
	"Accounts Manager", "Bộ phận đặt hàng",
}

# Bang gia mua cua ERPNext. Tao neu chua co - site nay dung ten mac dinh.
BANG_GIA_MUA = "Standard Buying"


def _kiem(quyen, viec):
	if not quyen & set(frappe.get_roles()):
		frappe.throw("Tài khoản của bạn không có quyền %s." % viec)


def _sua_duoc():
	return bool(QUYEN_SUA & set(frappe.get_roles()))


def _bao_dam_bang_gia():
	if frappe.db.exists("Price List", BANG_GIA_MUA):
		return BANG_GIA_MUA
	d = frappe.get_doc({
		"doctype": "Price List",
		"price_list_name": BANG_GIA_MUA,
		"buying": 1,
		"selling": 0,
		"enabled": 1,
		"currency": "VND",
	})
	d.insert(ignore_permissions=True)
	return d.name


def _bao_dam_dvt(ten):
	"""Don vi tinh. Tao neu chua co - Uyen go 'Túi' ma he chua co thi ket."""
	ten = (ten or "").strip()
	if not ten:
		return None
	if frappe.db.exists("UOM", ten):
		return ten
	d = frappe.get_doc({"doctype": "UOM", "uom_name": ten, "enabled": 1})
	d.insert(ignore_permissions=True)
	return d.name


def _gia_hien(ma_mon):
	r = frappe.get_all(
		"Item Price",
		filters={"item_code": ma_mon, "price_list": BANG_GIA_MUA, "buying": 1},
		fields=["name", "price_list_rate", "uom", "modified", "valid_from"],
		order_by="modified desc",
		limit_page_length=1,
	)
	return r[0] if r else None


def _quy_doi(ma_mon, dvt_mua):
	"""1 don vi mua bang bao nhieu don vi kho."""
	if not dvt_mua:
		return None
	r = frappe.get_all(
		"UOM Conversion Detail",
		filters={"parent": ma_mon, "uom": dvt_mua},
		fields=["conversion_factor"],
		limit_page_length=1,
	)
	return flt(r[0]["conversion_factor"]) if r else None


# ------------------------------------------------------------------ doc


@frappe.whitelist()
def danh_sach(tu_khoa="", chip=None, nhom=None, gioi_han=300):
	"""Mat hang mua kem gia mua dang khai."""
	_kiem(QUYEN_XEM, "xem bảng giá mua")
	_bao_dam_bang_gia()
	loc = {"is_purchase_item": 1, "disabled": 0}
	if (tu_khoa or "").strip():
		loc["item_name"] = ["like", "%" + tu_khoa.strip() + "%"]
	if nhom:
		loc["item_group"] = nhom
	mon = frappe.get_all(
		"Item",
		filters=loc,
		fields=["name", "item_name", "stock_uom", "item_group", "purchase_uom"],
		order_by="item_name asc",
		limit_page_length=0,
	)
	if not mon:
		return {"rows": [], "dem": {}, "tat_ca": 0, "sua_duoc": 1 if _sua_duoc() else 0}

	ma = [m["name"] for m in mon]
	gia = {}
	for g in frappe.get_all(
		"Item Price",
		filters={"item_code": ["in", ma], "price_list": BANG_GIA_MUA, "buying": 1},
		fields=["item_code", "price_list_rate", "uom", "modified"],
		order_by="modified asc",
		limit_page_length=0,
	):
		gia[g["item_code"]] = g
	qd = {}
	for c in frappe.get_all(
		"UOM Conversion Detail",
		filters={"parent": ["in", ma]},
		fields=["parent", "uom", "conversion_factor"],
		limit_page_length=0,
	):
		qd.setdefault(c["parent"], {})[c["uom"]] = flt(c["conversion_factor"])

	rows = []
	dem = {"co_gia": 0, "chua_gia": 0, "lech_dvt": 0}
	for m in mon:
		g = gia.get(m["name"]) or {}
		dvt_mua = g.get("uom") or m.get("purchase_uom") or m["stock_uom"]
		he_so = (qd.get(m["name"]) or {}).get(dvt_mua)
		if dvt_mua == m["stock_uom"]:
			he_so = 1.0
		don_gia = flt(g.get("price_list_rate"))
		gia_kho = (don_gia / he_so) if (don_gia and he_so) else 0.0
		r = {
			"ma": m["name"],
			"ten": m["item_name"],
			"nhom": m["item_group"],
			"dvt_kho": m["stock_uom"],
			"dvt_mua": dvt_mua,
			"he_so": he_so,
			"gia_mua": don_gia,
			"gia_kho": gia_kho,
			"sua_luc": str(g.get("modified") or ""),
		}
		if don_gia:
			dem["co_gia"] += 1
		else:
			dem["chua_gia"] += 1
		# Khai gia theo don vi mua khac don vi kho ma chua khai quy doi thi
		# may khong biet chia bao nhieu - do la dung cai bay lap xuong.
		if dvt_mua != m["stock_uom"] and not he_so:
			dem["lech_dvt"] += 1
			r["canh_bao"] = (
				"Đơn vị mua là %s mà đơn vị kho là %s, chưa khai quy đổi nên "
				"máy chưa tính được giá mỗi %s."
				% (dvt_mua, m["stock_uom"], m["stock_uom"])
			)
		rows.append(r)

	tat_ca = len(rows)
	if chip == "co_gia":
		rows = [r for r in rows if r["gia_mua"]]
	elif chip == "chua_gia":
		rows = [r for r in rows if not r["gia_mua"]]
	elif chip == "lech_dvt":
		rows = [r for r in rows if r.get("canh_bao")]

	return {
		"rows": rows[: cint(gioi_han) or 300],
		"dem": dem,
		"tat_ca": tat_ca,
		"con_nua": max(0, len(rows) - (cint(gioi_han) or 300)),
		"nhom": sorted({m["item_group"] for m in mon if m["item_group"]}),
		"sua_duoc": 1 if _sua_duoc() else 0,
	}


# ------------------------------------------------------------------ ghi


@frappe.whitelist()
def dat_gia(ma_mon, gia, dvt_mua=None, he_so=None):
	"""Khai gia mua cho mot mat hang.

	gia    - gia moi DON VI MUA (moi tui), khong phai moi don vi kho.
	he_so  - 1 don vi mua bang bao nhieu don vi kho (1 tui = 400 gram).
	"""
	_kiem(QUYEN_SUA, "khai giá mua")
	mon = frappe.db.get_value(
		"Item", ma_mon, ["name", "item_name", "stock_uom"], as_dict=True
	)
	if not mon:
		frappe.throw("Không thấy mặt hàng %s." % ma_mon)
	gia = flt(gia)
	if gia <= 0:
		frappe.throw("Giá mua phải lớn hơn 0.")

	dvt = _bao_dam_dvt(dvt_mua) or mon.stock_uom
	hs = flt(he_so) if he_so else (1.0 if dvt == mon.stock_uom else 0.0)
	if dvt != mon.stock_uom:
		if hs <= 0:
			frappe.throw(
				"Đơn vị mua là %s mà đơn vị kho là %s. Phải khai 1 %s bằng bao "
				"nhiêu %s, không thì máy không tính được giá vốn."
				% (dvt, mon.stock_uom, dvt, mon.stock_uom)
			)
		# Ghi quy doi len mat hang de don dat hang va phieu nhap dung theo.
		doc = frappe.get_doc("Item", ma_mon)
		co = [u for u in doc.uoms if u.uom == dvt]
		if co:
			co[0].conversion_factor = hs
		else:
			doc.append("uoms", {"uom": dvt, "conversion_factor": hs})
		doc.purchase_uom = dvt
		doc.flags.ignore_permissions = True
		doc.save(ignore_permissions=True)

	_bao_dam_bang_gia()
	cu = _gia_hien(ma_mon)
	if cu:
		p = frappe.get_doc("Item Price", cu["name"])
		p.price_list_rate = gia
		p.uom = dvt
		p.save(ignore_permissions=True)
	else:
		frappe.get_doc({
			"doctype": "Item Price",
			"item_code": ma_mon,
			"price_list": BANG_GIA_MUA,
			"buying": 1,
			"currency": "VND",
			"uom": dvt,
			"price_list_rate": gia,
			"valid_from": nowdate(),
		}).insert(ignore_permissions=True)

	gia_kho = gia / hs if hs else gia
	return {
		"ok": 1,
		"gia_kho": gia_kho,
		"loi_nhan": "Đã khai %s: %s đ mỗi %s%s."
		% (
			mon.item_name, "{:,.0f}".format(gia), dvt,
			", tức %s đ mỗi %s" % ("{:,.0f}".format(gia_kho), mon.stock_uom)
			if dvt != mon.stock_uom else "",
		),
	}


@frappe.whitelist()
def mau_excel():
	"""Tep mau de Uyen do ca bang gia vao mot lan."""
	_kiem(QUYEN_XEM, "tải mẫu bảng giá")
	from openpyxl import Workbook

	wb = Workbook()
	ws = wb.active
	ws.title = "Bang gia mua"
	ws.append(["Mã hàng", "Tên hàng", "Đơn vị kho", "Đơn vị mua", "1 đơn vị mua bằng bao nhiêu đơn vị kho", "Giá mỗi đơn vị mua"])
	for m in frappe.get_all(
		"Item",
		filters={"is_purchase_item": 1, "disabled": 0},
		fields=["name", "item_name", "stock_uom"],
		order_by="item_name asc",
		limit_page_length=0,
	):
		ws.append([m["name"], m["item_name"], m["stock_uom"], "", "", ""])
	for i, w in enumerate([16, 46, 13, 13, 34, 18], start=1):
		ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
	buf = io.BytesIO()
	wb.save(buf)
	return {
		"ten_file": "Bang-gia-mua-NVL.xlsx",
		"b64": base64.b64encode(buf.getvalue()).decode(),
		"kieu": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
	}


@frappe.whitelist()
def nhap_excel(b64=None, that_su=0):
	"""Doc tep Excel Uyen do len. that_su=0 chi soat, khong ghi gi.

	Luon soat truoc roi moi ghi: mot lan do nham cot la sai gia von hang
	loat, ma gia von sai thi khong nhin ra ngay - no chi lo ra o bao cao
	cuoi thang.
	"""
	_kiem(QUYEN_SUA, "nhập bảng giá")
	from openpyxl import load_workbook

	if not b64:
		frappe.throw("Chưa chọn tệp.")
	try:
		wb = load_workbook(io.BytesIO(base64.b64decode(b64)), data_only=True)
	except Exception:
		frappe.throw("Không đọc được tệp. Tệp phải là Excel .xlsx.")
	ws = wb.active

	ok, bo_qua, loi = [], 0, []
	for i, hang in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
		if not hang or not (hang[0] or ""):
			continue
		ma = str(hang[0]).strip()
		dvt_mua = str(hang[3] or "").strip() if len(hang) > 3 else ""
		hs = flt(hang[4]) if len(hang) > 4 and hang[4] not in (None, "") else 0.0
		gia = flt(hang[5]) if len(hang) > 5 and hang[5] not in (None, "") else 0.0
		if not gia:
			bo_qua += 1
			continue
		mon = frappe.db.get_value("Item", ma, ["name", "item_name", "stock_uom"], as_dict=True)
		if not mon:
			loi.append("Dòng %d: không có mã hàng %s" % (i, ma))
			continue
		dvt = dvt_mua or mon.stock_uom
		if dvt != mon.stock_uom and hs <= 0:
			loi.append(
				"Dòng %d (%s): đơn vị mua %s khác đơn vị kho %s mà chưa điền quy đổi"
				% (i, ma, dvt, mon.stock_uom)
			)
			continue
		ok.append({
			"ma": ma, "ten": mon.item_name, "dvt_kho": mon.stock_uom,
			"dvt_mua": dvt, "he_so": hs or 1.0, "gia": gia,
			"gia_kho": gia / (hs or 1.0),
		})

	if not cint(that_su):
		return {
			"thu": 1, "so_ok": len(ok), "so_loi": len(loi), "bo_qua": bo_qua,
			"loi": loi[:40], "mau": ok[:15],
		}

	da, hong = 0, []
	for r in ok:
		try:
			dat_gia(r["ma"], r["gia"], r["dvt_mua"], r["he_so"])
			da += 1
		except Exception as e:
			hong.append("%s: %s" % (r["ma"], str(e)[:120]))
	frappe.db.commit()
	return {
		"ok": 1, "da_khai": da, "so_loi": len(loi) + len(hong),
		"loi": (loi + hong)[:40],
		"loi_nhan": "Đã khai giá cho %d mặt hàng.%s"
		% (da, " %d dòng lỗi, xem danh sách bên dưới." % (len(loi) + len(hong))
		   if (loi or hong) else ""),
	}
